import requests
from bs4 import BeautifulSoup
from time import sleep
from random import randint
import openpyxl

def zhenshchin():
    book = openpyxl.Workbook()
    sheet = book.active

    sheet.cell(row=1, column=1).value = "Компания"
    sheet.cell(row=1, column=2).value = "Название"
    sheet.cell(row=1, column=3).value = "Принадлежность полу"
    sheet.cell(row=1, column=4).value = "Объём"
    sheet.cell(row=1, column=5).value = "Цена"
    sheet.cell(row=1, column=6).value = "Наличие"
    sheet.cell(row=1, column=7).value = "Код"
    sheet.cell(row=1, column=8).value = "Описание"
    sheet.cell(row=1, column=9).value = "Состав"
    sheet.cell(row=1, column=10).value = "Ссылки на фото"

    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/102.0.5005.115 Safari/537.36 OPR/88.0.4412.74 (Edition Campaign 34)"
        }

    spisok = []
    card_list = []
    rid = 2

    url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/"
    req = requests.get(url=url, headers=headers)
    soup = BeautifulSoup(req.text, "lxml")
    data = soup.find("div", class_="styled__StyledPagerWrapper-sc-qypphz-0").find_all("span")

    for i in data:
        spisok.append(i.text)
     
        
    for count in range(1, int(spisok[-1]) + 1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
            
        print(f"Итерация 1.{count} завершена")
        print()
        sleep(randint(2, 4))

    for count in range((int(spisok[-1])), 0, -1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 2.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1]) // 2), int(spisok[-1]) + 1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 3.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1]) // 2), 0, -1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 4.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(int(spisok[-1]), int(spisok[-1]) // 2):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 5.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(1, int(spisok[-1]) // 2):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 6.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(1, int(spisok[-1]) + 1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 7.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1])), 0, -1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 8.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(1, int(spisok[-1]) + 1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 9.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1])), 0, -1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 10.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1]) // 2), int(spisok[-1]) + 1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 11.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1]) // 2), 0, -1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 12.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(int(spisok[-1]), int(spisok[-1]) // 2):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 13.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(1, int(spisok[-1]) // 2):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 14.{count} завершена...")
        sleep(randint(2, 4))

    for count in range(1, int(spisok[-1]) + 1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 15.{count} завершена...")
        sleep(randint(2, 4))

    for count in range((int(spisok[-1])), 0, -1):
        url = f"https://www.notino.ua/parfyumeriya-dlya-zhenshchin/?f={count}-1-55544-55545"
        req = requests.get(url=url, headers=headers)
        soup = BeautifulSoup(req.text, "lxml")
        data = soup.find_all("div", class_="sc-gXmSlM")

        for i in data:
            try:
                card_url = "https://www.notino.ua" + i.find("a").get("href")
                card_list.append(card_url)
            except:
                pass
        print(f"Итерация 16.{count} завершена...")
        sleep(randint(2, 4))

    print("Ожидайте, ещё долго...")





    

    with open("card_url_list_#1.txt", "w", encoding="utf-8") as file:
        with open("card_url_list_#1.txt", "a", encoding="utf-8") as file:
            for line in card_list:
                file.write(f"{line}\n")

    with open("card_url_list_#1.txt", encoding="utf-8") as file:
        lines = [line.strip() for line in file.readlines()]

        lines1 = []
    
        for line in lines:
            lines1.append(line.strip())

        for i in range(len(lines1) - 1):
            for j in range(len(lines1) - i - 1):
                if(lines1[i] == lines1[j + i + 1]):
                    lines[j + i + 1] = ""

        for i in range(len(lines)):
            try:
                lines.remove("")
            except:
                pass

        lenght = len(lines)
        p1 = 0
        
        for line in lines:
            p2 = 0
            p1 += 1
            req = requests.get(url=line, headers=headers)
            soup = BeautifulSoup(req.text, "lxml")
            try:
                data = soup.find("ul", class_="sc-1hj58ih-6").find_all("li")

                for j in data:
                    p2 += 1
                    url = "https://www.notino.ua" + j.find("a").get("href")
                    req = requests.get(url=url, headers=headers)
                    soup = BeautifulSoup(req.text, "lxml")
                    
                    try:
                        full_desc = ""
                        full_src = ""
                        full_sostav = ""
                        
                        data = soup.find("div", class_="sc-ksluID")
                        company = data.find("div", class_="sc-3sotvb-0").find("span", class_="sc-3sotvb-3").find("a").text
                        name = data.find("div", class_="sc-3sotvb-0").find("span", class_="sc-3sotvb-3").find("span").text.strip()
                        pol = data.find("div", class_="sc-3sotvb-0").find("span", class_="sc-3sotvb-5").text
                        volume = data.find("div", class_="h83s98-4").text.strip()
                        try:
                            price = data.find("div", class_="mirfw-1").text.replace(".", "")
                        except:
                            price = "-"
                        
                        nal = data.find("div", class_="mu8uqe-0").find("span", class_="mu8uqe-4").text
                        kod = data.find("div", class_="mu8uqe-0").find("span", class_="mu8uqe-1").text
                                
                        try:
                            desc = data.find("div", {"id": "description-tab"}).find_all("p")
                        except:
                            desc = data.find("div", {"id": "description-tab"}).find("p")
                                    
                        lst = data.find("div", {"id": "description-tab"}).find("ul").find_all("li")
                        url_img = data.find_all("div", class_="slick-slide")
                        src = url_img[0].find("img").get("src").replace("detail_thumb", "detail_zoom")
                                        
                        for i in range(2):
                            if(i >= 1):
                                src = src.replace(f"_0{i}", f"_0{i + 1}").replace("detail_thumb", "detail_zoom")
                            full_src += src + "\n"

                        for i in range(len(desc)):
                            full_desc += desc[i].text + "\n"
                            if(i == 0):
                                for x in lst:
                                    full_desc += "-" + x.text + "\n"
                                full_desc += "\n"

                        sostav = data.find("div", {"id": "description-tab"}).find("div", class_="sc-1eu1dd2-5")
                        sostav = sostav.find("table", class_="sc-1eu1dd2-7").find_all("tr")
                                        
                        for i in sostav:
                            response = i.find_all("td")
                            for j in response:
                                full_sostav += j.text + "\n"
                            full_sostav += "\n"

                        full_desc = full_desc.replace("Склад аромату", "Склад аромату\n")

                        sheet.cell(row=rid, column=1).value = company
                        sheet.cell(row=rid, column=2).value = name
                        sheet.cell(row=rid, column=3).value = pol
                        sheet.cell(row=rid, column=4).value = volume
                        sheet.cell(row=rid, column=5).value = price
                        sheet.cell(row=rid, column=6).value = nal
                        sheet.cell(row=rid, column=7).value = kod
                        sheet.cell(row=rid, column=8).value = full_desc
                        sheet.cell(row=rid, column=9).value = full_sostav
                        sheet.cell(row=rid, column=10).value = full_src

                        rid += 1

                    except:
                        pass

                    print(f"Итерация {p1}.{p2} завершена...")
                    sleep(randint(2, 4))


            except:
                
                try:
                    full_desc = ""
                    full_src = ""
                    full_sostav = ""
                        
                    data = soup.find("div", class_="sc-ksluID")
                    company = data.find("div", class_="sc-3sotvb-0").find("span", class_="sc-3sotvb-3").find("a").text
                    name = data.find("div", class_="sc-3sotvb-0").find("span", class_="sc-3sotvb-3").find("span").text.strip()
                    pol = data.find("div", class_="sc-3sotvb-0").find("span", class_="sc-3sotvb-5").text
                    volume = data.find("div", class_="h83s98-4").text.strip()
                    try:
                        price = data.find("div", class_="mirfw-1").text.replace(".", "")
                    except:
                        price = "-"
                        
                    nal = data.find("div", class_="mu8uqe-0").find("span", class_="mu8uqe-4").text
                    kod = data.find("div", class_="mu8uqe-0").find("span", class_="mu8uqe-1").text
                                
                    try:
                        desc = data.find("div", {"id": "description-tab"}).find_all("p")
                    except:
                        desc = data.find("div", {"id": "description-tab"}).find("p")
                                    
                    lst = data.find("div", {"id": "description-tab"}).find("ul").find_all("li")
                    url_img = data.find_all("div", class_="slick-slide")
                    src = url_img[0].find("img").get("src").replace("detail_thumb", "detail_zoom")
                                        
                    for i in range(2):
                        if(i >= 1):
                            src = src.replace(f"_0{i}", f"_0{i + 1}").replace("detail_thumb", "detail_zoom")
                        full_src += src + "\n"

                    for i in range(len(desc)):
                        full_desc += desc[i].text + "\n"
                        if(i == 0):
                            for x in lst:
                                full_desc += "-" + x.text + "\n"
                            full_desc += "\n"

                    sostav = data.find("div", {"id": "description-tab"}).find("div", class_="sc-1eu1dd2-5")
                    sostav = sostav.find("table", class_="sc-1eu1dd2-7").find_all("tr")
                                        
                    for i in sostav:
                        response = i.find_all("td")
                        for j in response:
                            full_sostav += j.text + "\n"
                        full_sostav += "\n"

                    full_desc = full_desc.replace("Склад аромату", "Склад аромату\n")

                    sheet.cell(row=rid, column=1).value = company
                    sheet.cell(row=rid, column=2).value = name
                    sheet.cell(row=rid, column=3).value = pol
                    sheet.cell(row=rid, column=4).value = volume
                    sheet.cell(row=rid, column=5).value = price
                    sheet.cell(row=rid, column=6).value = nal
                    sheet.cell(row=rid, column=7).value = kod
                    sheet.cell(row=rid, column=8).value = full_desc
                    sheet.cell(row=rid, column=9).value = full_sostav
                    sheet.cell(row=rid, column=10).value = full_src

                    rid += 1

                except:
                    pass

                print(f"Итерация {p1}.1 завершена...")
                sleep(randint(2, 4))


                
                    
            lenght -= 1
            print(f"{lenght} итераций осталось.")
            
    book.save("parfum_table_#zhenshin.xlsx")
    book.close()

zhenshchin()

input("""Работа программы завершена.
Нажмите Enter...""")
